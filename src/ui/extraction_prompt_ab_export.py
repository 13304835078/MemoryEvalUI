from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.eval.extraction_prompt_compare import build_extraction_prompt_diff
from src.persistence import atomic_write_bytes


_BASE_COLUMNS = [
    "session_id",
    "chunk_id",
    "query",
    "answer",
    "评测人",
    "A提取结果",
    "B提取结果",
]
_REASONING_COLUMNS = ["A_reasoning", "B_reasoning"]
_COMMON_COMPARISON_COLUMNS = [
    "源数据一致性",
    "A提取状态",
    "B提取状态",
    "对比结论",
    "对比备注",
    "A错误标签",
    "B错误标签",
]
_LEGACY_COMPARISON_COLUMNS = [
    "A Judge状态",
    "B Judge状态",
    "A总分",
    "B总分",
    "B-A",
    "A维度得分",
    "B维度得分",
    "A评语",
    "B评语",
    "A规则引用",
    "B规则引用",
]
_DIRECT_COMPARISON_COLUMNS = [
    "对比调用状态",
    "对比模型",
    "对比置信度",
    "判定依据类型",
    "规则引用",
    "策略差异",
    "证据引用",
    "A相对问题",
    "B相对问题",
    "A相对优点",
    "B相对优点",
    "对比调用错误",
]
_EXCEL_CELL_LIMIT = 32_767


def _excel_safe(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    if len(cleaned) <= _EXCEL_CELL_LIMIT:
        return cleaned
    suffix = "\n[内容超过 Excel 单元格上限，已截断；完整内容见原始提取文件]"
    return cleaned[: _EXCEL_CELL_LIMIT - len(suffix)] + suffix


def _dataframe(rows: list[dict[str, Any]], *, include_reasoning: bool) -> pd.DataFrame:
    direct_mode = any(str(row.get("对比调用状态") or "").strip() for row in rows)
    mode_columns = _DIRECT_COMPARISON_COLUMNS if direct_mode else _LEGACY_COMPARISON_COLUMNS
    columns = (
        _BASE_COLUMNS
        + (_REASONING_COLUMNS if include_reasoning else [])
        + _COMMON_COMPARISON_COLUMNS
        + mode_columns
    )
    safe_rows = [
        {column: _excel_safe(row.get(column, "")) for column in columns}
        for row in rows
    ]
    return pd.DataFrame(safe_rows, columns=columns)


def _column_width(header: str) -> int:
    if header in {"session_id", "chunk_id"}:
        return 14
    if header == "评测人":
        return 18
    if header in {"query", "answer"}:
        return 42
    if "提取结果" in header or "reasoning" in header:
        return 68
    if "评语" in header or "备注" in header or "规则引用" in header or "策略差异" in header or "证据引用" in header or "相对" in header:
        return 48
    if "维度得分" in header or "错误标签" in header:
        return 34
    if "状态" in header or header in {"对比结论", "源数据一致性"}:
        return 18
    return 14


def _normalized_cell(value: Any) -> str:
    return "\n".join(
        line.rstrip()
        for line in str(value or "").replace("\r\n", "\n").split("\n")
    ).strip()


def _style_diff_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    header_by_name: dict[str, int] = {}
    default_header_fill = PatternFill("solid", fgColor="A61B1B")
    a_header_fill = PatternFill("solid", fgColor="4472C4")
    b_header_fill = PatternFill("solid", fgColor="C65911")
    for index, cell in enumerate(worksheet[1], start=1):
        header = str(cell.value or "")
        header_by_name[header] = index
        cell.fill = a_header_fill if header.startswith("A") else (
            b_header_fill if header.startswith("B") else default_header_fill
        )
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.column_dimensions[get_column_letter(index)].width = _column_width(header)
    worksheet.row_dimensions[1].height = 28

    wrapped_headers = {
        "query",
        "answer",
        "A提取结果",
        "B提取结果",
        "A_reasoning",
        "B_reasoning",
        "A维度得分",
        "B维度得分",
        "对比备注",
        "A错误标签",
        "B错误标签",
        "A评语",
        "B评语",
        "A规则引用",
        "B规则引用",
        "规则引用",
        "策略差异",
        "证据引用",
        "A相对问题",
        "B相对问题",
        "A相对优点",
        "B相对优点",
        "对比调用错误",
    }
    for header in wrapped_headers:
        column = header_by_name.get(header)
        if column is None:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    changed_a_fill = PatternFill("solid", fgColor="DDEBF7")
    changed_b_fill = PatternFill("solid", fgColor="FCE4D6")
    same_fill = PatternFill("solid", fgColor="E7E6E6")
    for left_header, right_header in (
        ("A提取结果", "B提取结果"),
        ("A_reasoning", "B_reasoning"),
    ):
        left_column = header_by_name.get(left_header)
        right_column = header_by_name.get(right_header)
        if left_column is None or right_column is None:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            left_cell = worksheet.cell(row=row_index, column=left_column)
            right_cell = worksheet.cell(row=row_index, column=right_column)
            left = _normalized_cell(left_cell.value)
            right = _normalized_cell(right_cell.value)
            if not left and not right:
                continue
            if left == right:
                left_cell.fill = same_fill
                right_cell.fill = same_fill
            else:
                left_cell.fill = changed_a_fill
                right_cell.fill = changed_b_fill

    comparison_column = header_by_name.get("对比结论")
    if comparison_column is not None:
        conclusion_fills = {
            "A较优": PatternFill("solid", fgColor="FFF2CC"),
            "B较优": PatternFill("solid", fgColor="E2F0D9"),
            "输出相同": same_fill,
            "基本持平": same_fill,
            "双方均漏抽": same_fill,
            "不可比较": PatternFill("solid", fgColor="FCE4D6"),
            "策略差异": PatternFill("solid", fgColor="E4DFEC"),
        }
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=comparison_column)
            fill = conclusion_fills.get(str(cell.value or ""))
            if fill is not None:
                cell.fill = fill


def write_extraction_prompt_diff_excel(
    *,
    extraction_a_path: str | Path,
    extraction_b_path: str | Path,
    comparison_rows: list[dict[str, Any]],
    output_path: str | Path,
    model_comparison: dict[str, Any] | None = None,
) -> Path:
    extraction_a_path = Path(extraction_a_path)
    extraction_b_path = Path(extraction_b_path)
    output_path = Path(output_path)
    rows_a = pd.read_excel(extraction_a_path).fillna("").to_dict("records")
    rows_b = pd.read_excel(extraction_b_path).fillna("").to_dict("records")
    row_diff, chunk_diff, include_reasoning = build_extraction_prompt_diff(
        rows_a,
        rows_b,
        comparison_rows,
    )

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _dataframe(row_diff, include_reasoning=include_reasoning).to_excel(
            writer,
            sheet_name="逐行Diff",
            index=False,
        )
        _dataframe(chunk_diff, include_reasoning=include_reasoning).to_excel(
            writer,
            sheet_name="逐Chunk对比",
            index=False,
        )
        pd.DataFrame(
            [
                {"项目": "逐行Diff", "说明": "保留原始 query/answer 行；chunk 级输出与结论只写在末行。"},
                {"项目": "逐Chunk对比", "说明": "每个 session/chunk 一行，便于筛选两版结果。"},
                {"项目": "reasoning", "说明": "仅当至少一版返回 reasoning 时加入 A/B reasoning 列。"},
                {"项目": "失败样本", "说明": "API、网络和 Judge 解析失败单独标注，不按 0 分计入。"},
                {"项目": "策略差异", "说明": "由两版准入范围、数据源或输出结构不同造成，不进入 A/B 胜负统计。"},
            ]
        ).to_excel(writer, sheet_name="说明", index=False)
        if model_comparison:
            model_row = {
                key: "；".join(str(item) for item in value) if isinstance(value, list) else _excel_safe(value)
                for key, value in model_comparison.items()
            }
            pd.DataFrame([model_row]).to_excel(writer, sheet_name="模型综合意见", index=False)
        _style_diff_sheet(writer.book["逐行Diff"])
        _style_diff_sheet(writer.book["逐Chunk对比"])
        explanation = writer.book["说明"]
        explanation.freeze_panes = "A2"
        explanation.column_dimensions["A"].width = 18
        explanation.column_dimensions["B"].width = 88
        for cell in explanation[1]:
            cell.fill = PatternFill("solid", fgColor="A61B1B")
            cell.font = Font(color="FFFFFF", bold=True)
        for row in explanation.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if "模型综合意见" in writer.book.sheetnames:
            model_sheet = writer.book["模型综合意见"]
            model_sheet.freeze_panes = "A2"
            for index, cell in enumerate(model_sheet[1], start=1):
                cell.fill = PatternFill("solid", fgColor="A61B1B")
                cell.font = Font(color="FFFFFF", bold=True)
                model_sheet.column_dimensions[get_column_letter(index)].width = 42
            for row in model_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    atomic_write_bytes(output_path, buffer.getvalue())
    return output_path
